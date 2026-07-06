"""Karoo Online Excel — an insight-led workbook. Leads with an executive
summary and an Insights & Solutions sheet (finding -> so-what -> action),
then the supporting data and charts."""
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "out"
CHARTS = OUT / "charts"
INK, CREAM, OCHRE, GOLD, TEAL = "2B2118", "FBF3EC", "C0531F", "E0A526", "2A7F7F"

tk = json.load(open(OUT / "tk_summary.json"))
rb = json.load(open(OUT / "rb_summary.json"))
mk = json.load(open(OUT / "mkt_summary.json"))
lc = json.load(open(OUT / "lifecycle_summary.json"))
oil = json.load(open(OUT / "oil_summary.json"))

# ── Executive summary (narrative KPIs) ──
exec_rows = [
    ("CATALOGUE", "", ""),
    ("Products analysed (catalogue)", f"{tk['n_products']:,}", "R25–R53,451, median R492"),
    ("In stock for fast delivery", f"{tk['pct_in_stock']}%", f"{tk['pct_lead_time']}% imported lead-time"),
    ("DEMAND", "", ""),
    ("Metro share of revenue", "72%", "3 provinces; on-time 72% vs interior 32%"),
    ("MARKETING", "", ""),
    ("Blended paid ROAS", f"{mk['paid_roas']:.1f}x", f"best: {mk['best_paid_channel']} {mk['best_paid_roas']:.1f}x"),
    ("Remarketing ROAS", f"{mk['remarketing_roas']:.1f}x", f"{mk['best_remarketing']} {mk['best_remarketing_roas']:.0f}x"),
    ("AI share of voice", f"{mk['ai_share_of_voice']}%", "vs Takealot 41%, Amazon 23%"),
    ("CUSTOMERS", "", ""),
    ("Average lifetime value", f"R{lc['avg_clv']:,}", f"top 10% = {lc['champions_clv']}% of CLV"),
    ("Churn rate", f"{lc['churn_rate_pct']}%", f"M1 {int(lc['m1_retention'])}% -> M6 {int(lc['m6_retention'])}% retention"),
    ("At-risk customers", f"{lc['at_risk_customers']:,}", f"R{lc['at_risk_revenue']}m of revenue slipping"),
    ("RISK & RESILIENCE", "", ""),
    ("Fraud review-queue precision", f"{int(rb['review_queue_fraud']/rb['review_queue']*100)}%", f"catches {int(rb['review_queue_fraud']/rb['true_fraud']*100)}% of fraud"),
    ("2026 oil-crisis margin hit", f"R{oil['margin_hit']/1e6:.1f}m", f"Brent peak ${oil['peak_brent']} (+{oil['peak_shock_pct']}%), Mar 2026, {oil['worst_category']} worst"),
]
exec_df = pd.DataFrame(exec_rows, columns=["Metric", "Value", "Context"])

# ── Insights & Solutions (finding -> so-what -> action) ──
insights = pd.DataFrame([
    ["Availability", f"{tk['pct_lead_time']}% of products ship on 4-16 day import lead times",
     "Availability disappoints more than price; slow items hurt satisfaction",
     "Source top-selling lines locally; surface stock status in search & PLP"],
    ["Delivery", "Metros deliver 72% on-time in 2.9 days; interior 32% in 4.7 days",
     "Geography, not courier, drives speed and returns (corr -0.63)",
     "Add a regional fulfilment hub to cut interior delivery time and returns"],
    ["Marketing", f"Remarketing returns {mk['remarketing_roas']:.0f}x vs {mk['paid_roas']:.1f}x on cold paid",
     "Recovering existing intent is far cheaper than buying new clicks",
     f"Shift budget to {mk['best_remarketing']} & cart-recovery flows"],
    ["Recommendations", f"{rb['top_rule']} bought together at {rb['top_lift']:.2f}x lift",
     "Proven affinities lift basket size at zero acquisition cost",
     "Add 'customers also bought' widgets on PLP + cart for top pairs"],
    ["Churn / CLV", f"Top 10% of customers drive {lc['champions_clv']}% of lifetime value; {lc['at_risk_customers']:,} at risk",
     "A small, high-value cohort is slipping and is cheap to save",
     "Launch a targeted winback for At-Risk high-CLV customers"],
    ["AI visibility", f"{mk['ai_share_of_voice']}% share of voice in AI answers vs Takealot 41%",
     "Customers increasingly ask AI assistants, not just Google",
     "Structured data, FAQ content & reviews to grow AI citations"],
    ["Fraud", f"2-signal review queue = {rb['review_queue']:,} accounts at {int(rb['review_queue_fraud']/rb['review_queue']*100)}% precision",
     "No single rule works, but stacked signals are highly precise",
     "Route the queue to a daily manual-review workflow"],
    ["Resilience", f"2026 Hormuz crisis (Brent +{oil['peak_shock_pct']}%) cost ~R{oil['margin_hit']/1e6:.1f}m over Feb-Jun 2026",
     f"{oil['worst_category']} & discretionary demand are exposed to fuel inflation",
     "Hedge fuel contracts; promote resilient essentials during spikes"],
], columns=["Area", "Finding", "So what", "Recommended action"])

sheets = {
    "Executive Summary": exec_df,
    "Insights & Solutions": insights,
    "Product Catalogue": pd.read_csv(OUT / "takealot_products_clean.csv"),
    "Category Summary": pd.read_csv(OUT / "tk_category_summary.csv"),
    "Province Stats": pd.read_csv(OUT / "province_stats.csv"),
    "Customer Lifecycle": pd.read_csv(OUT / "lifecycle_stages.csv"),
    "CLV Bands": pd.read_csv(OUT / "clv_bands.csv"),
    "Recommendations": pd.read_csv(OUT / "rb_recommendations.csv"),
    "Correlations": pd.read_csv(OUT / "rb_correlation_matrix.csv"),
    "Fraud Signals": pd.read_csv(OUT / "rb_fraud_summary.csv"),
    "Paid Campaigns": pd.read_csv(OUT / "mkt_paid_by_channel.csv"),
    "Remarketing": pd.read_csv(OUT / "mkt_remarketing.csv"),
    "AI Visibility": pd.read_csv(OUT / "mkt_ai_visibility.csv"),
    "Oil Impact (monthly)": pd.read_csv(OUT / "oil_impact_monthly.csv"),
}

_site = Path(r"C:\Users\Anthony.DESKTOP-ES5HL78\Downloads\karoo-online-site\reports")
xlsx = OUT / "Karoo_Online_Data_Story.xlsx"
try:
    with open(xlsx, "a"):
        pass
except PermissionError:
    _site.mkdir(parents=True, exist_ok=True)
    xlsx = _site / "Karoo_Online_Data_Story.xlsx"
    print(f"(out/ copy locked -> writing to site: {xlsx})")

with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
    for name, df in sheets.items():
        df.to_excel(xw, sheet_name=name[:31], index=(name == "Correlations"))

wb = load_workbook(xlsx)
thin = Side(style="thin", color="E5D8CC")
for ws in wb.worksheets:
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.font = Font(color=CREAM, bold=True, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, 12), 60)
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        for c in row:
            c.border = Border(bottom=thin)
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=CREAM)

# style Executive Summary section-header rows (Value empty)
es = wb["Executive Summary"]
for row in es.iter_rows(min_row=2):
    if row[1].value in ("", None) and row[0].value:
        for c in row:
            c.fill = PatternFill("solid", fgColor=OCHRE)
            c.font = Font(color="FFFFFF", bold=True, size=10)
# wrap Insights sheet
ins = wb["Insights & Solutions"]
for row in ins.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical="top")
for col_letter in ["B", "C", "D"]:
    ins.column_dimensions[col_letter].width = 42

ws = wb.create_sheet("Charts")
ws["A1"] = "Karoo Online — Key Visuals"
ws["A1"].font = Font(color=OCHRE, bold=True, size=14)
anchor = 3
for chart in ["data_provenance.png", "tk_price_dist.png", "sa_choropleth.png",
              "lifecycle_stages.png", "cohort_retention.png", "clv_bands.png",
              "mkt_paid.png", "mkt_ai_sov.png", "oil_delivery.png", "oil_revenue.png",
              "rb_recommendations.png", "rb_correlation.png", "rb_fraud.png"]:
    p = CHARTS / chart
    if p.exists():
        im = XLImage(str(p)); sc = 640 / im.width
        im.width = int(im.width * sc); im.height = int(im.height * sc)
        ws.add_image(im, f"A{anchor}"); anchor += int(im.height / 18) + 3
wb.save(xlsx)
print(f"Excel: {xlsx} | {len(wb.sheetnames)} sheets")
