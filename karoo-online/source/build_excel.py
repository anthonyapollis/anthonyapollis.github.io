"""Build the Karoo Data Story Excel workbook — full product/brand/category/
province detail, formatted for analysis, with the embedded charts."""
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUT = HERE / "out"
CHARTS = OUT / "charts"

products = pd.read_csv(OUT / "products_master.csv")
brands = pd.read_csv(OUT / "brands.csv")
cats = pd.read_csv(OUT / "category_summary.csv")
cat_totals = pd.read_csv(OUT / "category_totals.csv")
prov_cat = pd.read_csv(OUT / "province_category.csv")

prov = prov_cat.groupby("province", as_index=False)["revenue"].sum().sort_values("revenue", ascending=False)

xlsx = OUT / "Karoo_Data_Story.xlsx"

OCHRE = "C0531F"
INK = "2B2118"
CREAM = "FBF3EC"

sheets = {
    "Executive Summary": None,  # built manually
    "All Products": products,
    "Top 10 Products": products.head(10),
    "Bottom 10 Products": products.tail(10).iloc[::-1],
    "Brands": brands,
    "Categories": cats,
    "Category (source data)": cat_totals,
    "Provinces": prov,
}

with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
    # Executive summary sheet
    summary = pd.DataFrame({
        "Metric": ["Total revenue (R)", "Total units sold", "Products", "Brands", "Categories",
                   "Avg product rating", "Top product", "Top brand", "Top category", "Top province"],
        "Value": [
            f"{products['revenue'].sum():,.0f}", f"{products['units_sold'].sum():,}",
            len(products), brands['brand'].nunique(), len(cats),
            f"{products['avg_rating'].mean():.2f}",
            f"{products.iloc[0]['product_name']} ({products.iloc[0]['brand']})",
            brands.iloc[0]['brand'], cats.iloc[0]['category'], prov.iloc[0]['province'],
        ],
    })
    summary.to_excel(xw, sheet_name="Executive Summary", index=False)
    for name, df in sheets.items():
        if df is not None:
            df.to_excel(xw, sheet_name=name[:31], index=False)

# ── formatting pass ──
wb = load_workbook(xlsx)
header_fill = PatternFill("solid", fgColor=INK)
header_font = Font(color=CREAM, bold=True, size=10)
title_font = Font(color=OCHRE, bold=True, size=14)

for ws in wb.worksheets:
    # header row styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    # auto column widths
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 3, 11), 42)
    # zebra striping
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if i % 2 == 1:
            for c in row:
                c.fill = PatternFill("solid", fgColor=CREAM)

# add a Charts sheet with the key visuals
ws = wb.create_sheet("Charts")
ws["A1"] = "Karoo Data Story — Key Visuals"
ws["A1"].font = title_font
row_anchor = 3
for chart in ["category_treemap.png", "top10_products.png", "sa_choropleth.png",
              "top_brands.png", "return_vs_revenue.png"]:
    p = CHARTS / chart
    if p.exists():
        img = XLImage(str(p))
        # scale down to ~640px wide
        scale = 640 / img.width
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, f"A{row_anchor}")
        row_anchor += int(img.height / 18) + 3

wb.save(xlsx)
print("Excel workbook:", xlsx)
print("Sheets:", wb.sheetnames)
